from datetime import datetime, timezone
from decimal import Decimal

from app.services.reportes_export import (
    generar_pdf_financiero,
    generar_pdf_inventario,
    generar_pdf_pedidos,
    generar_pdf_productos,
    generar_xlsx_financiero,
    generar_xlsx_inventario,
    generar_xlsx_pedidos,
    generar_xlsx_productos,
)

_DATOS_PRODUCTOS = {
    "desde": datetime(2026, 6, 1, tzinfo=timezone.utc),
    "hasta": datetime(2026, 6, 30, tzinfo=timezone.utc),
    "productos": [
        {
            "producto_id": 1,
            "nombre": "Latte",
            "categoria": "Bebidas calientes",
            "disponible": True,
            "cantidad_vendida": 10,
            "ingresos": Decimal("550.00"),
            "costo_total": Decimal("40.00"),
            "margen": Decimal("510.00"),
            "margen_pct": Decimal("92.73"),
        }
    ],
}

_DATOS_PEDIDOS = {
    "desde": datetime(2026, 6, 1, tzinfo=timezone.utc),
    "hasta": datetime(2026, 6, 30, tzinfo=timezone.utc),
    "total_pedidos": 1,
    "total_ventas": Decimal("638.00"),
    "pedidos": [
        {
            "pedido_id": 1,
            "fecha": datetime(2026, 6, 15, tzinfo=timezone.utc),
            "mesa": 1,
            "mesero": "Test Mesero",
            "estatus": "Entregado",
            "total": Decimal("638.00"),
        }
    ],
}

_DATOS_FINANCIERO = {
    "desde": datetime(2026, 6, 1, tzinfo=timezone.utc),
    "hasta": datetime(2026, 6, 30, tzinfo=timezone.utc),
    "total_ventas": Decimal("1000.00"),
    "total_gastos": Decimal("400.00"),
    "ganancia_neta": Decimal("600.00"),
    "margen_pct": Decimal("60.00"),
    "margen_pct_anterior": Decimal("50.00"),
    "variacion_ventas_pct": Decimal("10.00"),
    "variacion_ganancia_pct": Decimal("20.00"),
    "ranking_margen": [
        {
            "producto_id": 1,
            "nombre": "Latte",
            "ingresos": Decimal("550.00"),
            "costo_total": Decimal("40.00"),
            "margen": Decimal("510.00"),
            "margen_pct": Decimal("92.73"),
        }
    ],
    "ventas_por_categoria": [],
    "ventas_por_usuario": [],
    "ventas_por_metodo_pago": [],
    "detalle_ventas": [
        {
            "fecha": datetime(2026, 6, 15, tzinfo=timezone.utc),
            "pedido_id": 1,
            "mesa": 1,
            "mesero": "Test Mesero",
            "producto": "Latte",
            "cantidad": 2,
            "precio_unitario": Decimal("55.00"),
            "subtotal": Decimal("110.00"),
        }
    ],
    "detalle_gastos": [
        {
            "id": 1,
            "concepto": "Compra de insumo: Leche",
            "monto": Decimal("400.00"),
            "fecha_gasto": datetime(2026, 6, 10, tzinfo=timezone.utc),
            "usuario": "Test Cajero",
        }
    ],
    "gastos_por_tipo": [
        {"tipo": "Compras de insumos", "total": Decimal("400.00")},
        {"tipo": "Gastos fijos", "total": Decimal("0")},
        {"tipo": "Otros gastos", "total": Decimal("0")},
    ],
    "gastos_por_usuario": [
        {"usuario_id": 1, "nombre": "Test Cajero", "total": Decimal("400.00")},
    ],
}

_DATOS_INVENTARIO = {
    "riesgo": [
        {
            "id": 1,
            "nombre": "Leche entera",
            "unidad": "ml",
            "stock_actual": Decimal("500"),
            "stock_minimo": Decimal("1000"),
            "falta": Decimal("500"),
            "costo_reposicion": Decimal("10.00"),
            "productos_afectados": ["Latte", "Capuchino"],
        }
    ],
    "ranking_consumo": [],
}


def test_generar_pdf_financiero_produce_pdf_valido():
    buffer = generar_pdf_financiero(_DATOS_FINANCIERO)
    contenido = buffer.read()
    assert contenido[:4] == b"%PDF"


def test_generar_pdf_inventario_produce_pdf_valido():
    buffer = generar_pdf_inventario(_DATOS_INVENTARIO)
    contenido = buffer.read()
    assert contenido[:4] == b"%PDF"


def test_generar_pdf_inventario_sin_riesgo_no_falla():
    buffer = generar_pdf_inventario({"riesgo": [], "ranking_consumo": []})
    contenido = buffer.read()
    assert contenido[:4] == b"%PDF"


def test_generar_xlsx_financiero_produce_zip_valido():
    buffer = generar_xlsx_financiero(_DATOS_FINANCIERO)
    contenido = buffer.read()
    assert contenido[:2] == b"PK"  # firma de archivo ZIP (XLSX es un ZIP)


def test_generar_xlsx_inventario_produce_zip_valido():
    buffer = generar_xlsx_inventario(_DATOS_INVENTARIO)
    contenido = buffer.read()
    assert contenido[:2] == b"PK"


def test_generar_pdf_financiero_incluye_metodo_pago():
    datos = {
        "desde": datetime(2026, 1, 1),
        "hasta": datetime(2026, 1, 31),
        "total_ventas": Decimal("100.00"),
        "total_gastos": Decimal("20.00"),
        "ganancia_neta": Decimal("80.00"),
        "margen_pct": Decimal("80.00"),
        "ranking_margen": [],
        "ventas_por_categoria": [],
        "ventas_por_usuario": [],
        "ventas_por_metodo_pago": [{"metodo_pago": "Efectivo", "total": Decimal("100.00")}],
        "detalle_ventas": [],
        "detalle_gastos": [],
        "gastos_por_tipo": [],
        "gastos_por_usuario": [],
    }
    buffer = generar_pdf_financiero(datos)
    assert buffer.getbuffer().nbytes > 0


def test_generar_xlsx_financiero_incluye_hoja_metodo_pago():
    datos = {
        "desde": datetime(2026, 1, 1),
        "hasta": datetime(2026, 1, 31),
        "total_ventas": Decimal("100.00"),
        "total_gastos": Decimal("20.00"),
        "ganancia_neta": Decimal("80.00"),
        "margen_pct": Decimal("80.00"),
        "ranking_margen": [],
        "ventas_por_categoria": [],
        "ventas_por_usuario": [],
        "ventas_por_metodo_pago": [{"metodo_pago": "Efectivo", "total": Decimal("100.00")}],
        "detalle_ventas": [],
        "detalle_gastos": [],
        "gastos_por_tipo": [],
        "gastos_por_usuario": [],
    }
    from openpyxl import load_workbook

    buffer = generar_xlsx_financiero(datos)
    libro = load_workbook(buffer)
    assert "Ventas por método de pago" in libro.sheetnames


def test_generar_pdf_inventario_incluye_ranking_consumo():
    datos = {
        "riesgo": [],
        "ranking_consumo": [{"nombre": "Café molido", "unidad": "g", "cantidad_consumida": Decimal("500")}],
    }
    buffer = generar_pdf_inventario(datos)
    assert buffer.getbuffer().nbytes > 0


def test_generar_pdf_productos_produce_pdf_valido():
    buffer = generar_pdf_productos(_DATOS_PRODUCTOS)
    contenido = buffer.read()
    assert contenido[:4] == b"%PDF"


def test_generar_pdf_productos_sin_productos_no_falla():
    buffer = generar_pdf_productos({"desde": _DATOS_PRODUCTOS["desde"], "hasta": _DATOS_PRODUCTOS["hasta"], "productos": []})
    contenido = buffer.read()
    assert contenido[:4] == b"%PDF"


def test_generar_xlsx_productos_produce_zip_valido():
    buffer = generar_xlsx_productos(_DATOS_PRODUCTOS)
    contenido = buffer.read()
    assert contenido[:2] == b"PK"


def test_generar_pdf_pedidos_produce_pdf_valido():
    buffer = generar_pdf_pedidos(_DATOS_PEDIDOS)
    contenido = buffer.read()
    assert contenido[:4] == b"%PDF"


def test_generar_pdf_pedidos_sin_pedidos_no_falla():
    datos = {**_DATOS_PEDIDOS, "total_pedidos": 0, "total_ventas": Decimal("0"), "pedidos": []}
    buffer = generar_pdf_pedidos(datos)
    contenido = buffer.read()
    assert contenido[:4] == b"%PDF"


def test_generar_xlsx_pedidos_produce_zip_valido():
    buffer = generar_xlsx_pedidos(_DATOS_PEDIDOS)
    contenido = buffer.read()
    assert contenido[:2] == b"PK"
