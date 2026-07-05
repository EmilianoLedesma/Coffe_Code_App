import io
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

_HOUSE = colors.HexColor("#1E3932")
_ACCENT = colors.HexColor("#00754A")
_LIGHT = colors.HexColor("#F2F0EB")
_GREY = colors.HexColor("#6B7280")

_ESTILO_TABLA = TableStyle(
    [
        ("BACKGROUND", (0, 0), (-1, 0), _HOUSE),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 9),
        ("FONTSIZE", (0, 1), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [_LIGHT, colors.white]),
        ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]
)

_RELLENO_ENCABEZADO = PatternFill(start_color="1E3932", end_color="1E3932", fill_type="solid")
_FUENTE_ENCABEZADO = Font(color="FFFFFF", bold=True)


def _documento_base(buffer: io.BytesIO) -> SimpleDocTemplate:
    return SimpleDocTemplate(
        buffer,
        pagesize=A4,
        leftMargin=2 * cm,
        rightMargin=2 * cm,
        topMargin=2 * cm,
        bottomMargin=2 * cm,
    )


def _estilos_parrafo():
    base = getSampleStyleSheet()
    h1 = ParagraphStyle("h1", parent=base["Heading1"], textColor=_ACCENT, fontSize=18, spaceAfter=4)
    h2 = ParagraphStyle("h2", parent=base["Heading2"], textColor=_HOUSE, fontSize=11, spaceBefore=10, spaceAfter=4)
    sub = ParagraphStyle("sub", parent=base["Normal"], textColor=_GREY, fontSize=9, spaceAfter=8)
    ftr = ParagraphStyle("ftr", parent=base["Normal"], textColor=colors.grey, fontSize=8)
    return h1, h2, sub, ftr


def _tabla(filas: list[list[str]]) -> Table:
    tabla = Table(filas, hAlign="LEFT")
    tabla.setStyle(_ESTILO_TABLA)
    return tabla


SECCIONES_FINANCIERO = {
    "resumen",
    "rendimiento",
    "categoria",
    "usuario",
    "metodo_pago",
    "detalle_ventas",
    "gastos_tipo",
    "gastos_usuario",
    "detalle_gastos",
}
SECCIONES_INVENTARIO = {"riesgo", "consumo"}


def generar_pdf_financiero(datos: dict, secciones: set[str] | None = None) -> io.BytesIO:
    incluir = lambda clave: secciones is None or clave in secciones  # noqa: E731

    buffer = io.BytesIO()
    doc = _documento_base(buffer)
    h1, h2, sub, ftr = _estilos_parrafo()

    story = [Paragraph("Coffee Code — Reporte Financiero", h1)]
    story.append(
        Paragraph(f"Periodo: {datos['desde'].strftime('%d/%m/%Y')} — {datos['hasta'].strftime('%d/%m/%Y')}", sub)
    )

    if incluir("resumen"):
        story.append(
            _tabla(
                [
                    ["Ventas", "Gastos", "Ganancia neta", "Margen %"],
                    [
                        f"${datos['total_ventas']:,.2f}",
                        f"${datos['total_gastos']:,.2f}",
                        f"${datos['ganancia_neta']:,.2f}",
                        f"{datos['margen_pct']}%",
                    ],
                ]
            )
        )

    if incluir("rendimiento") and datos["ranking_margen"]:
        story.append(Paragraph("Rendimiento de producto", h2))
        story.append(
            _tabla(
                [["Producto", "Ingresos", "Costo estimado", "Margen", "Margen %"]]
                + [
                    [
                        fila["nombre"],
                        f"${fila['ingresos']:,.2f}",
                        f"${fila['costo_total']:,.2f}",
                        f"${fila['margen']:,.2f}",
                        f"{fila['margen_pct']}%",
                    ]
                    for fila in datos["ranking_margen"]
                ]
            )
        )

    if incluir("categoria") and datos["ventas_por_categoria"]:
        story.append(Paragraph("Ventas por categoría", h2))
        story.append(
            _tabla(
                [["Categoría", "Total"]]
                + [[fila["nombre"], f"${fila['total']:,.2f}"] for fila in datos["ventas_por_categoria"]]
            )
        )

    if incluir("usuario") and datos["ventas_por_usuario"]:
        story.append(Paragraph("Ventas por mesero/cajero", h2))
        story.append(
            _tabla(
                [["Usuario", "Total"]]
                + [[fila["nombre"], f"${fila['total']:,.2f}"] for fila in datos["ventas_por_usuario"]]
            )
        )

    if incluir("metodo_pago") and datos["ventas_por_metodo_pago"]:
        story.append(Paragraph("Ventas por método de pago", h2))
        story.append(
            _tabla(
                [["Método de pago", "Total"]]
                + [[fila["metodo_pago"], f"${fila['total']:,.2f}"] for fila in datos["ventas_por_metodo_pago"]]
            )
        )

    if incluir("detalle_ventas") and datos["detalle_ventas"]:
        story.append(Paragraph("Detalle de ventas", h2))
        story.append(
            _tabla(
                [["Fecha", "Pedido", "Mesa", "Mesero", "Producto", "Cant.", "Subtotal"]]
                + [
                    [
                        fila["fecha"].strftime("%d/%m/%Y %H:%M"),
                        f"#{fila['pedido_id']}",
                        str(fila["mesa"]),
                        fila["mesero"],
                        fila["producto"],
                        str(fila["cantidad"]),
                        f"${fila['subtotal']:,.2f}",
                    ]
                    for fila in datos["detalle_ventas"]
                ]
            )
        )

    if incluir("gastos_tipo") and datos["gastos_por_tipo"]:
        story.append(Paragraph("Gastos por tipo", h2))
        story.append(
            _tabla(
                [["Tipo", "Total"]]
                + [[fila["tipo"], f"${fila['total']:,.2f}"] for fila in datos["gastos_por_tipo"]]
            )
        )

    if incluir("gastos_usuario") and datos["gastos_por_usuario"]:
        story.append(Paragraph("Gastos por usuario", h2))
        story.append(
            _tabla(
                [["Usuario", "Total"]]
                + [[fila["nombre"], f"${fila['total']:,.2f}"] for fila in datos["gastos_por_usuario"]]
            )
        )

    if incluir("detalle_gastos") and datos["detalle_gastos"]:
        story.append(Paragraph("Detalle de gastos", h2))
        story.append(
            _tabla(
                [["Fecha", "Concepto", "Usuario", "Monto"]]
                + [
                    [
                        fila["fecha_gasto"].strftime("%d/%m/%Y %H:%M"),
                        fila["concepto"],
                        fila["usuario"],
                        f"${fila['monto']:,.2f}",
                    ]
                    for fila in datos["detalle_gastos"]
                ]
            )
        )

    story.append(Spacer(1, 0.8 * cm))
    story.append(Paragraph(f"Generado por Coffee Code API · {datetime.now().strftime('%d/%m/%Y %H:%M')}", ftr))
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_pdf_inventario(datos: dict, secciones: set[str] | None = None) -> io.BytesIO:
    incluir = lambda clave: secciones is None or clave in secciones  # noqa: E731

    buffer = io.BytesIO()
    doc = _documento_base(buffer)
    h1, h2, sub, ftr = _estilos_parrafo()

    story = [Paragraph("Coffee Code — Reporte de Inventario", h1)]
    story.append(Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", sub))

    if incluir("riesgo"):
        if datos["riesgo"]:
            story.append(Paragraph("Riesgo de inventario", h2))
            story.append(
                _tabla(
                    [["Ingrediente", "Falta", "Costo de reposición", "Productos afectados"]]
                    + [
                        [
                            fila["nombre"],
                            f"{fila['falta']} {fila['unidad']}",
                            f"${fila['costo_reposicion']:,.2f}",
                            ", ".join(fila["productos_afectados"]),
                        ]
                        for fila in datos["riesgo"]
                    ]
                )
            )
        else:
            story.append(Paragraph("Sin ingredientes bajo el stock mínimo.", sub))

    if incluir("consumo") and datos["ranking_consumo"]:
        story.append(Paragraph("Ranking de consumo de ingredientes", h2))
        story.append(
            _tabla(
                [["Ingrediente", "Cantidad consumida"]]
                + [
                    [fila["nombre"], f"{fila['cantidad_consumida']} {fila['unidad']}"]
                    for fila in datos["ranking_consumo"]
                ]
            )
        )

    story.append(Spacer(1, 0.8 * cm))
    story.append(Paragraph(f"Generado por Coffee Code API · {datetime.now().strftime('%d/%m/%Y %H:%M')}", ftr))
    doc.build(story)
    buffer.seek(0)
    return buffer


def _libro_vacio() -> Workbook:
    libro = Workbook()
    libro.remove(libro.active)
    return libro


def _agregar_hoja(libro: Workbook, nombre: str, encabezados: list[str]):
    hoja = libro.create_sheet(nombre)
    hoja.append(encabezados)
    for celda in hoja[1]:
        celda.fill = _RELLENO_ENCABEZADO
        celda.font = _FUENTE_ENCABEZADO
    return hoja


def generar_xlsx_financiero(datos: dict, secciones: set[str] | None = None) -> io.BytesIO:
    incluir = lambda clave: secciones is None or clave in secciones  # noqa: E731
    libro = _libro_vacio()

    if incluir("resumen"):
        hoja_resumen = _agregar_hoja(libro, "Resumen financiero", ["Métrica", "Valor"])
        hoja_resumen.append(["Ventas", float(datos["total_ventas"])])
        hoja_resumen.append(["Gastos", float(datos["total_gastos"])])
        hoja_resumen.append(["Ganancia neta", float(datos["ganancia_neta"])])
        hoja_resumen.append(["Margen %", float(datos["margen_pct"])])

    if incluir("rendimiento"):
        hoja_ranking = _agregar_hoja(
            libro, "Rendimiento de producto", ["Producto", "Ingresos", "Costo estimado", "Margen", "Margen %"]
        )
        for fila in datos["ranking_margen"]:
            hoja_ranking.append(
                [
                    fila["nombre"],
                    float(fila["ingresos"]),
                    float(fila["costo_total"]),
                    float(fila["margen"]),
                    float(fila["margen_pct"]),
                ]
            )

    if incluir("categoria"):
        hoja_categoria = _agregar_hoja(libro, "Ventas por categoría", ["Categoría", "Total"])
        for fila in datos["ventas_por_categoria"]:
            hoja_categoria.append([fila["nombre"], float(fila["total"])])

    if incluir("usuario"):
        hoja_usuario = _agregar_hoja(libro, "Ventas por usuario", ["Usuario", "Total"])
        for fila in datos["ventas_por_usuario"]:
            hoja_usuario.append([fila["nombre"], float(fila["total"])])

    if incluir("metodo_pago"):
        hoja_metodo_pago = _agregar_hoja(libro, "Ventas por método de pago", ["Método de pago", "Total"])
        for fila in datos["ventas_por_metodo_pago"]:
            hoja_metodo_pago.append([fila["metodo_pago"], float(fila["total"])])

    if incluir("detalle_ventas"):
        hoja_detalle_ventas = _agregar_hoja(
            libro, "Detalle de ventas", ["Fecha", "Pedido", "Mesa", "Mesero", "Producto", "Cantidad", "Subtotal"]
        )
        for fila in datos["detalle_ventas"]:
            hoja_detalle_ventas.append(
                [
                    fila["fecha"].strftime("%d/%m/%Y %H:%M"),
                    fila["pedido_id"],
                    fila["mesa"],
                    fila["mesero"],
                    fila["producto"],
                    fila["cantidad"],
                    float(fila["subtotal"]),
                ]
            )

    if incluir("gastos_tipo"):
        hoja_gastos_tipo = _agregar_hoja(libro, "Gastos por tipo", ["Tipo", "Total"])
        for fila in datos["gastos_por_tipo"]:
            hoja_gastos_tipo.append([fila["tipo"], float(fila["total"])])

    if incluir("gastos_usuario"):
        hoja_gastos_usuario = _agregar_hoja(libro, "Gastos por usuario", ["Usuario", "Total"])
        for fila in datos["gastos_por_usuario"]:
            hoja_gastos_usuario.append([fila["nombre"], float(fila["total"])])

    if incluir("detalle_gastos"):
        hoja_detalle_gastos = _agregar_hoja(
            libro, "Detalle de gastos", ["Fecha", "Concepto", "Usuario", "Monto"]
        )
        for fila in datos["detalle_gastos"]:
            hoja_detalle_gastos.append(
                [
                    fila["fecha_gasto"].strftime("%d/%m/%Y %H:%M"),
                    fila["concepto"],
                    fila["usuario"],
                    float(fila["monto"]),
                ]
            )

    if not libro.sheetnames:
        _agregar_hoja(libro, "Reporte financiero", ["Sin secciones seleccionadas"])

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer


def generar_xlsx_inventario(datos: dict, secciones: set[str] | None = None) -> io.BytesIO:
    incluir = lambda clave: secciones is None or clave in secciones  # noqa: E731
    libro = _libro_vacio()

    if incluir("riesgo"):
        hoja = _agregar_hoja(
            libro, "Riesgo de inventario", ["Ingrediente", "Falta", "Unidad", "Costo de reposición", "Productos afectados"]
        )
        for fila in datos["riesgo"]:
            hoja.append(
                [
                    fila["nombre"],
                    float(fila["falta"]),
                    fila["unidad"],
                    float(fila["costo_reposicion"]),
                    ", ".join(fila["productos_afectados"]),
                ]
            )

    if incluir("consumo"):
        hoja_consumo = _agregar_hoja(libro, "Ranking de consumo", ["Ingrediente", "Unidad", "Cantidad consumida"])
        for fila in datos["ranking_consumo"]:
            hoja_consumo.append([fila["nombre"], fila["unidad"], float(fila["cantidad_consumida"])])

    if not libro.sheetnames:
        _agregar_hoja(libro, "Reporte de inventario", ["Sin secciones seleccionadas"])

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer


def generar_pdf_productos(datos: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = _documento_base(buffer)
    h1, h2, sub, ftr = _estilos_parrafo()

    story = [Paragraph("Coffee Code — Reporte de Productos", h1)]
    story.append(
        Paragraph(f"Periodo: {datos['desde'].strftime('%d/%m/%Y')} — {datos['hasta'].strftime('%d/%m/%Y')}", sub)
    )

    if datos["productos"]:
        story.append(Paragraph("Catálogo y desempeño", h2))
        story.append(
            _tabla(
                [["Producto", "Categoría", "Disponible", "Vendidos", "Ingresos", "Margen", "Margen %"]]
                + [
                    [
                        fila["nombre"],
                        fila["categoria"],
                        "Sí" if fila["disponible"] else "No",
                        str(fila["cantidad_vendida"]),
                        f"${fila['ingresos']:,.2f}",
                        f"${fila['margen']:,.2f}",
                        f"{fila['margen_pct']}%",
                    ]
                    for fila in datos["productos"]
                ]
            )
        )
    else:
        story.append(Paragraph("Sin productos activos registrados.", sub))

    story.append(Spacer(1, 0.8 * cm))
    story.append(Paragraph(f"Generado por Coffee Code API · {datetime.now().strftime('%d/%m/%Y %H:%M')}", ftr))
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_xlsx_productos(datos: dict) -> io.BytesIO:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Catálogo de productos"
    hoja.append(["Producto", "Categoría", "Disponible", "Vendidos", "Ingresos", "Costo estimado", "Margen", "Margen %"])
    for celda in hoja[1]:
        celda.fill = _RELLENO_ENCABEZADO
        celda.font = _FUENTE_ENCABEZADO
    for fila in datos["productos"]:
        hoja.append(
            [
                fila["nombre"],
                fila["categoria"],
                "Sí" if fila["disponible"] else "No",
                fila["cantidad_vendida"],
                float(fila["ingresos"]),
                float(fila["costo_total"]),
                float(fila["margen"]),
                float(fila["margen_pct"]),
            ]
        )

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer


def generar_pdf_pedidos(datos: dict) -> io.BytesIO:
    buffer = io.BytesIO()
    doc = _documento_base(buffer)
    h1, h2, sub, ftr = _estilos_parrafo()

    story = [Paragraph("Coffee Code — Reporte de Pedidos", h1)]
    story.append(
        Paragraph(f"Periodo: {datos['desde'].strftime('%d/%m/%Y')} — {datos['hasta'].strftime('%d/%m/%Y')}", sub)
    )
    story.append(
        _tabla(
            [
                ["Total de pedidos", "Total ventas"],
                [str(datos["total_pedidos"]), f"${datos['total_ventas']:,.2f}"],
            ]
        )
    )

    if datos["pedidos"]:
        story.append(Paragraph("Detalle de pedidos", h2))
        story.append(
            _tabla(
                [["Pedido", "Fecha", "Mesa", "Mesero", "Estatus", "Total"]]
                + [
                    [
                        f"#{fila['pedido_id']}",
                        fila["fecha"].strftime("%d/%m/%Y %H:%M"),
                        str(fila["mesa"]),
                        fila["mesero"],
                        fila["estatus"],
                        f"${fila['total']:,.2f}",
                    ]
                    for fila in datos["pedidos"]
                ]
            )
        )
    else:
        story.append(Paragraph("Sin pedidos en el periodo seleccionado.", sub))

    story.append(Spacer(1, 0.8 * cm))
    story.append(Paragraph(f"Generado por Coffee Code API · {datetime.now().strftime('%d/%m/%Y %H:%M')}", ftr))
    doc.build(story)
    buffer.seek(0)
    return buffer


def generar_xlsx_pedidos(datos: dict) -> io.BytesIO:
    libro = Workbook()
    hoja = libro.active
    hoja.title = "Pedidos"
    hoja.append(["Pedido", "Fecha", "Mesa", "Mesero", "Estatus", "Total"])
    for celda in hoja[1]:
        celda.fill = _RELLENO_ENCABEZADO
        celda.font = _FUENTE_ENCABEZADO
    for fila in datos["pedidos"]:
        hoja.append(
            [
                fila["pedido_id"],
                fila["fecha"].strftime("%d/%m/%Y %H:%M"),
                fila["mesa"],
                fila["mesero"],
                fila["estatus"],
                float(fila["total"]),
            ]
        )

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)
    return buffer
