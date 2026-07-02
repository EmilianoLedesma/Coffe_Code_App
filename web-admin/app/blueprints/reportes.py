import io
from datetime import date, datetime, timedelta

from flask import Blueprint, Response, render_template, request
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from weasyprint import HTML

from app.api_client import (
    listar_ingredientes,
    listar_productos,
    listar_receta_producto,
    obtener_reporte_admin,
)
from app.auth import api_base_url, current_token, login_required
from app.reportes import (
    calcular_margen_pct,
    costo_receta,
    mapa_ingrediente_a_productos,
    ranking_margen,
    riesgo_inventario,
)

bp = Blueprint("reportes", __name__, url_prefix="/reportes")


def _parsear_fecha(valor: str | None, default: date) -> date:
    if not valor:
        return default
    return datetime.strptime(valor, "%Y-%m-%d").date()


def _construir_datos_reporte(desde: date, hasta: date) -> dict:
    token = current_token()
    base_url = api_base_url()

    reporte = obtener_reporte_admin(base_url, token, desde.isoformat(), hasta.isoformat())
    margen = calcular_margen_pct(reporte["total_ventas"], reporte["ganancia_neta"])

    top_productos = reporte["top_productos"]
    costos_por_producto = {}
    for producto in top_productos:
        receta = listar_receta_producto(base_url, token, producto["producto_id"])
        costos_por_producto[producto["producto_id"]] = costo_receta(receta) if receta else 0
    ranking = ranking_margen(top_productos, costos_por_producto)

    productos = listar_productos(base_url, token)
    productos_por_id = {p["id"]: p for p in productos}
    recetas_por_producto = {p["id"]: listar_receta_producto(base_url, token, p["id"]) for p in productos}
    mapa = mapa_ingrediente_a_productos(recetas_por_producto, productos_por_id)

    ingredientes = listar_ingredientes(base_url, token)
    riesgo = riesgo_inventario(ingredientes, mapa)

    return {"desde": desde, "hasta": hasta, "reporte": reporte, "margen": margen, "ranking": ranking, "riesgo": riesgo}


@bp.route("/exportar.pdf")
@login_required
def exportar_pdf():
    hoy = date.today()
    hasta = _parsear_fecha(request.args.get("hasta"), hoy)
    desde = _parsear_fecha(request.args.get("desde"), hoy - timedelta(days=30))

    datos = _construir_datos_reporte(desde, hasta)
    html_renderizado = render_template("reportes/reporte_pdf.html", **datos)
    pdf_bytes = HTML(string=html_renderizado).write_pdf()

    return Response(
        pdf_bytes,
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=reporte_{desde}_a_{hasta}.pdf"},
    )


@bp.route("/exportar.xlsx")
@login_required
def exportar_xlsx():
    hoy = date.today()
    hasta = _parsear_fecha(request.args.get("hasta"), hoy)
    desde = _parsear_fecha(request.args.get("desde"), hoy - timedelta(days=30))

    datos = _construir_datos_reporte(desde, hasta)

    libro = Workbook()
    encabezado_relleno = PatternFill(start_color="6F4E37", end_color="6F4E37", fill_type="solid")
    encabezado_fuente = Font(color="F5E6D3", bold=True)

    hoja_resumen = libro.active
    hoja_resumen.title = "Resumen financiero"
    hoja_resumen.append(["Métrica", "Valor"])
    for celda in hoja_resumen[1]:
        celda.fill = encabezado_relleno
        celda.font = encabezado_fuente
    hoja_resumen.append(["Ventas", float(datos["reporte"]["total_ventas"])])
    hoja_resumen.append(["Gastos", float(datos["reporte"]["total_gastos"])])
    hoja_resumen.append(["Ganancia neta", float(datos["reporte"]["ganancia_neta"])])
    hoja_resumen.append(["Margen %", float(datos["margen"])])

    hoja_ranking = libro.create_sheet("Rendimiento de producto")
    hoja_ranking.append(["Producto", "Ingresos", "Costo estimado", "Margen", "Margen %"])
    for celda in hoja_ranking[1]:
        celda.fill = encabezado_relleno
        celda.font = encabezado_fuente
    for fila in datos["ranking"]:
        hoja_ranking.append(
            [fila["nombre"], float(fila["ingresos"]), float(fila["costo_total"]), float(fila["margen"]), float(fila["margen_pct"])]
        )

    hoja_riesgo = libro.create_sheet("Riesgo de inventario")
    hoja_riesgo.append(["Ingrediente", "Falta", "Unidad", "Costo de reposición", "Productos afectados"])
    for celda in hoja_riesgo[1]:
        celda.fill = encabezado_relleno
        celda.font = encabezado_fuente
    for fila in datos["riesgo"]:
        hoja_riesgo.append(
            [fila["nombre"], float(fila["falta"]), fila["unidad"], float(fila["costo_reposicion"]), ", ".join(fila["productos_afectados"])]
        )

    buffer = io.BytesIO()
    libro.save(buffer)
    buffer.seek(0)

    return Response(
        buffer.read(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=reporte_{desde}_a_{hasta}.xlsx"},
    )
